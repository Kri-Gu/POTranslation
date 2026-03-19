"""
Streamlit frontend for the PO File Translator.

Run with:
    streamlit run src/app.py
"""

import os
import sys
import tempfile
import csv
import io
from pathlib import Path
from typing import Dict, List

import streamlit as st
from openpyxl import load_workbook

# Ensure the src package is importable
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.po_translate_en_to_nb import (
    AVAILABLE_MODELS,
    TARGET_LANGUAGES,
    SOURCE_LANGUAGES,
    translate_po_file,
    build_work_items,
    load_context,
    get_defaults,
    format_glossary_for_prompt,
)
from src.xliff_translate import (
    translate_xliff_file,
    build_work_items_xliff,
)
from src.cost_estimator import estimate_cost
import xml.etree.ElementTree as ET
import polib

# ─── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PO File Translator",
    page_icon="🌍",
    layout="wide",
)

# ─── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { max-width: 1200px; margin: 0 auto; }
    div[data-testid="stMetric"] {
        background-color: #f0f2f6;
        border-radius: 8px;
        padding: 12px 16px;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 8px;
        padding: 16px;
        margin: 8px 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeeba;
        border-radius: 8px;
        padding: 16px;
        margin: 8px 0;
    }
</style>
""", unsafe_allow_html=True)


def _dedupe_glossary(entries: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Keep first occurrence per source term and drop invalid rows."""
    deduped: List[Dict[str, str]] = []
    seen = set()
    for item in entries:
        src = str(item.get("source", "")).strip()
        tgt = str(item.get("target", "")).strip()
        if not src or not tgt:
            continue
        key = src.casefold()
        if key in seen:
            continue
        seen.add(key)
        deduped.append({"source": src, "target": tgt})
    return deduped


def _parse_glossary_text(text: str) -> List[Dict[str, str]]:
    """Parse glossary lines entered as 'source -> target' style rows."""
    parsed: List[Dict[str, str]] = []
    if not text.strip():
        return parsed

    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        for sep in ["→", "=", "->"]:
            if sep in line:
                src, tgt = line.split(sep, 1)
                parsed.append({"source": src.strip(), "target": tgt.strip()})
                break
    return _dedupe_glossary(parsed)


def _parse_glossary_upload(uploaded_file) -> List[Dict[str, str]]:
    """Parse glossary rows from .csv or .xlsx files."""
    if uploaded_file is None:
        return []

    name = uploaded_file.name.lower()
    parsed: List[Dict[str, str]] = []

    if name.endswith(".csv"):
        raw = uploaded_file.getvalue().decode("utf-8-sig", errors="replace")
        stream = io.StringIO(raw)
        reader = csv.DictReader(stream)
        field_map = {str(k).strip().lower(): k for k in (reader.fieldnames or [])}

        if "source" in field_map and "target" in field_map:
            for row in reader:
                parsed.append(
                    {
                        "source": str(row.get(field_map["source"], "")).strip(),
                        "target": str(row.get(field_map["target"], "")).strip(),
                    }
                )
        else:
            stream.seek(0)
            plain_reader = csv.reader(stream)
            for row in plain_reader:
                if len(row) < 2:
                    continue
                parsed.append({"source": str(row[0]).strip(), "target": str(row[1]).strip()})

    elif name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        first_row = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        has_headers = "source" in first_row and "target" in first_row

        if has_headers:
            src_idx = first_row.index("source")
            tgt_idx = first_row.index("target")
            data_rows = rows[1:]
        else:
            src_idx = 0
            tgt_idx = 1
            data_rows = rows

        for row in data_rows:
            if row is None:
                continue
            src = str(row[src_idx]).strip() if len(row) > src_idx and row[src_idx] is not None else ""
            tgt = str(row[tgt_idx]).strip() if len(row) > tgt_idx and row[tgt_idx] is not None else ""
            parsed.append({"source": src, "target": tgt})

    return _dedupe_glossary(parsed)


def main():
    st.title("🌍 Translation File Translator")
    st.caption("Translate .po (Poedit) and .xlf (XLIFF) files using OpenAI — preserving placeholders, inline markup, and formatting.")

    # Load user defaults from .env
    defaults = get_defaults()

    # ─── Sidebar: Settings ─────────────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Settings")

        # API key
        api_key = st.text_input(
            "OpenAI API Key",
            type="password",
            value=os.getenv("OPENAI_API_KEY", ""),
            help="Your OpenAI API key. Also reads from OPENAI_API_KEY env var.",
        )
        if api_key:
            os.environ["OPENAI_API_KEY"] = api_key

        st.divider()

        # Model selection — default from .env
        default_model_idx = AVAILABLE_MODELS.index(defaults["model"]) if defaults["model"] in AVAILABLE_MODELS else 0
        model = st.selectbox(
            "Model",
            options=AVAILABLE_MODELS,
            index=default_model_idx,
            help="gpt-4.1 recommended for quality. gpt-5.2 for best quality. gpt-4.1-mini/nano for lower cost.",
        )

        # Target language — default from .env
        target_keys = list(TARGET_LANGUAGES.keys())
        default_target_idx = target_keys.index(defaults["target_lang"]) if defaults["target_lang"] in target_keys else 0
        target_lang = st.selectbox(
            "Target language",
            options=target_keys,
            format_func=lambda k: f"{TARGET_LANGUAGES[k]} ({k})",
            index=default_target_idx,
        )

        # Source language — default from .env
        default_source_idx = SOURCE_LANGUAGES.index(defaults["source_lang"]) if defaults["source_lang"] in SOURCE_LANGUAGES else 0
        source_lang = st.selectbox(
            "Source language detection",
            options=SOURCE_LANGUAGES,
            index=default_source_idx,
            help="'auto' detects English vs German per entry. Use 'en' or 'de' to force.",
        )

        # Batch size — default from .env
        batch_size = st.slider(
            "Batch size",
            min_value=5,
            max_value=50,
            value=min(max(defaults["batch_size"], 5), 50),
            step=5,
            help="Items per API call. Smaller = better quality, larger = faster.",
        )

        # Force translate
        force = st.checkbox(
            "Force translate all entries",
            value=True,
            help="Translate every entry, even if it already has a translation.",
        )

        st.divider()

        # Context file
        st.subheader("📋 Domain context")
        context_mode = st.radio(
            "Context source",
            ["None", "Upload file", "Paste text"],
            horizontal=True,
        )
        context_text = ""
        if context_mode == "Upload file":
            ctx_file = st.file_uploader("Context file", type=["json", "txt", "md"])
            if ctx_file:
                context_text = ctx_file.read().decode("utf-8")
                st.success(f"Loaded {len(context_text)} chars")
        elif context_mode == "Paste text":
            context_text = st.text_area(
                "Domain instructions",
                height=120,
                placeholder="e.g. Translate technical language related to lawn and garden machinery…",
            )

        st.divider()

        # ── Glossary / Term protection ─────────────────────────────────────
        st.subheader("📖 Glossary")
        st.caption("Terms that must be translated exactly as specified.")

        # Initialise session state for glossary
        if "glossary" not in st.session_state:
            st.session_state.glossary = []

        glossary_input = st.text_area(
            "Glossary entries (one per line: source → target)",
            height=120,
            placeholder="lawn mower → gressklipper\nblade speed → knivhastighet\nthrottle → gasspedal",
            help="Enter one term pair per line, separated by → or =. These terms will be enforced during translation.",
        )

        glossary_file = st.file_uploader(
            "Upload glossary file (.csv or .xlsx)",
            type=["csv", "xlsx"],
            key="glossary_upload",
            help="CSV/XLSX with columns `source` and `target` (or first two columns as source/target).",
        )

        # Parse glossary from text input and optional file upload
        glossary_text_entries = _parse_glossary_text(glossary_input)
        glossary_file_entries = _parse_glossary_upload(glossary_file) if glossary_file else []
        glossary: list = _dedupe_glossary(glossary_text_entries + glossary_file_entries)

        st.session_state.glossary = glossary

        if glossary:
            st.success(f"{len(glossary)} glossary term(s) loaded")
            if glossary_file_entries:
                st.caption(f"{len(glossary_file_entries)} loaded from file upload")

        st.divider()
        st.caption("CLI equivalent shown after translation.")

    # ─── Main area: File upload & action ───────────────────────────────────
    col_upload, col_info = st.columns([2, 1])

    with col_upload:
        uploaded_file = st.file_uploader(
            "Upload a .po or .xlf file",
            type=["po", "xlf"],
            help="Upload a .po (Poedit/gettext) or .xlf (XLIFF 1.2) file to translate.",
        )

    # Detect file type and route accordingly
    if uploaded_file is not None:
        file_ext = Path(uploaded_file.name).suffix.lower()
        is_xliff = file_ext in {".xlf", ".xliff"}

        # ── Save upload to a temp file ──────────────────────────────────────
        suffix = ".xlf" if is_xliff else ".po"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix, mode="wb") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_input = tmp.name

        # ── Parse and build work items ──────────────────────────────────────
        if is_xliff:
            try:
                xliff_tree = ET.parse(tmp_input)
            except ET.ParseError as e:
                st.error(f"Failed to parse XLIFF file: {e}")
                return

            work_items, id_map, total_entries = build_work_items_xliff(
                xliff_tree, source_lang=source_lang, force=force
            )

            with col_info:
                st.metric("Trans-units", total_entries)
                st.metric("To translate", len(work_items))
                # XLIFF items with markup need smaller batches, cap display
                eff_batch = min(batch_size, 10)
                estimated_calls = (len(work_items) + eff_batch - 1) // eff_batch if eff_batch else 0
                st.metric("API calls (est.)", estimated_calls)

                # Cost estimation
                if work_items:
                    cost_est = estimate_cost(work_items, model, eff_batch)
                    st.metric("Est. tokens", f"{cost_est['estimated_total_tokens']:,}")
                    st.metric("Est. cost", f"${cost_est['estimated_cost_usd']:.4f}")

            with st.expander("📄 Preview trans-units to translate", expanded=False):
                preview_data = []
                for item in work_items[:100]:
                    preview_data.append({
                        "Has markup": "Yes" if item["has_markup"] else "No",
                        "Source text": item["text"][:150],
                    })
                st.dataframe(preview_data, use_container_width=True, hide_index=True)
                if len(work_items) > 100:
                    st.caption(f"Showing first 100 of {len(work_items)} trans-units.")

            st.divider()
            if not api_key:
                st.warning("⚠️ Enter your OpenAI API key in the sidebar to translate.")
                return

            col_btn, col_cli = st.columns([1, 2])
            with col_btn:
                translate_btn = st.button(
                    "🚀 Translate",
                    type="primary",
                    use_container_width=True,
                    disabled=(len(work_items) == 0),
                )
            with col_cli:
                cli_cmd = (
                    f'python src/xliff_translate.py "{uploaded_file.name}" "output.xlf"'
                    f" --model {model} --batch-size {min(batch_size, 10)}"
                    f" --target-lang {target_lang} --source-lang {source_lang}"
                )
                if force:
                    cli_cmd += " --force"
                if context_text:
                    cli_cmd += ' --context-file "context.json"'
                st.code(cli_cmd, language="powershell")

            if translate_btn:
                _run_xliff_translation(
                    tmp_input, model, batch_size, target_lang, source_lang,
                    force, context_text, uploaded_file.name, glossary,
                )

        else:
            # ── PO file flow (unchanged) ────────────────────────────────────
            try:
                po = polib.pofile(tmp_input, encoding="utf-8")
            except Exception as e:
                st.error(f"Failed to parse PO file: {e}")
                return

            work_items, id_map, total_entries = build_work_items(
                po, source_lang=source_lang, force=force
            )

            with col_info:
                st.metric("Total entries", total_entries)
                st.metric("To translate", len(work_items))
                estimated_calls = (len(work_items) + batch_size - 1) // batch_size
                st.metric("API calls (est.)", estimated_calls)

                # Cost estimation
                if work_items:
                    cost_est = estimate_cost(work_items, model, batch_size)
                    st.metric("Est. tokens", f"{cost_est['estimated_total_tokens']:,}")
                    st.metric("Est. cost", f"${cost_est['estimated_cost_usd']:.4f}")

            with st.expander("📄 Preview entries to translate", expanded=False):
                preview_data = []
                for item in work_items[:100]:
                    entry_obj, src_field = id_map[item["id"]]
                    preview_data.append({
                        "Source lang": item["lang"],
                        "Source field": src_field,
                        "Text": item["text"][:120],
                        "Current msgstr": entry_obj.msgstr[:120] if entry_obj.msgstr else "—",
                    })
                st.dataframe(preview_data, use_container_width=True, hide_index=True)
                if len(work_items) > 100:
                    st.caption(f"Showing first 100 of {len(work_items)} entries.")

            st.divider()
            if not api_key:
                st.warning("⚠️ Enter your OpenAI API key in the sidebar to translate.")
                return

            col_btn, col_cli = st.columns([1, 2])
            with col_btn:
                translate_btn = st.button(
                    "🚀 Translate",
                    type="primary",
                    use_container_width=True,
                    disabled=(len(work_items) == 0),
                )
            with col_cli:
                cli_cmd = (
                    f'python src/po_translate_en_to_nb.py "{uploaded_file.name}" "output.po"'
                    f" --model {model} --batch-size {batch_size}"
                    f" --target-lang {target_lang} --source-lang {source_lang}"
                )
                if force:
                    cli_cmd += " --force"
                if context_text:
                    cli_cmd += ' --context-file "context.json"'
                st.code(cli_cmd, language="powershell")

            if translate_btn:
                _run_translation(
                    tmp_input, model, batch_size, target_lang, source_lang,
                    force, context_text, uploaded_file.name, glossary,
                )

    else:
        st.info("👆 Upload a `.po` or `.xlf` file to get started.")

        with st.expander("ℹ️ How it works"):
            st.markdown("""
**Supported formats**
- `.po` — Poedit / gettext translation files
- `.xlf` — XLIFF 1.2 files (e-learning, CMS, product content)

**Workflow**
1. **Upload** your translation file.
2. **Configure** model, target language, and batch size in the sidebar.
3. **Optionally** add domain context for better terminology.
4. Click **Translate** and wait for the progress bar to complete.
5. **Download** the translated file.

Placeholders, HTML tags, and XLIFF inline codes (`<g>` elements) are preserved automatically.
            """)


def _run_translation(
    tmp_input, model, batch_size, target_lang, source_lang,
    force, context_text, original_filename, glossary=None,
):
    """Execute the translation and display results."""

    # Write context to a temp file if provided
    context_file = None
    if context_text:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".txt", mode="w", encoding="utf-8") as ctx_tmp:
            ctx_tmp.write(context_text)
            context_file = ctx_tmp.name

    # Create temp output path
    with tempfile.NamedTemporaryFile(delete=False, suffix=".po") as out_tmp:
        tmp_output = out_tmp.name

    # Progress bar and log area
    progress_bar = st.progress(0, text="Starting translation…")
    log_container = st.empty()
    log_messages = []

    def on_progress(translated, total):
        pct = translated / total if total > 0 else 1.0
        progress_bar.progress(pct, text=f"Translated {translated} / {total} entries")

    def on_log(msg):
        log_messages.append(msg)
        log_container.text("\n".join(log_messages[-5:]))

    # Run
    try:
        result = translate_po_file(
            tmp_input,
            tmp_output,
            model=model,
            batch_size=batch_size,
            target_lang=target_lang,
            source_lang=source_lang,
            force=force,
            context_file=context_file,
            progress_callback=on_progress,
            log_callback=on_log,
            glossary=glossary,
        )
    except Exception as e:
        st.error(f"❌ Translation failed: {e}")
        return

    progress_bar.progress(1.0, text="✅ Translation complete!")

    # ─── Results ───────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📊 Results")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("✅ Translated", result["translated"])
    col2.metric("📄 Total entries", result["total_entries"])
    col3.metric("⚠️ Placeholder issues", len(result["placeholder_warnings"]))
    col4.metric("❌ Failed", len(result["failed"]))

    # Placeholder warnings
    if result["placeholder_warnings"]:
        with st.expander(f"⚠️ {len(result['placeholder_warnings'])} placeholder warning(s)", expanded=True):
            pw_data = []
            for pw in result["placeholder_warnings"]:
                pw_data.append({
                    "ID": pw["id"],
                    "Missing": ", ".join(pw["missing"]),
                    "Source": pw["source"][:80],
                    "Translation": pw["translation"][:80],
                })
            st.dataframe(pw_data, use_container_width=True, hide_index=True)

    # Failed items
    if result["failed"]:
        with st.expander(f"❌ {len(result['failed'])} failed item(s)"):
            for fi in result["failed"]:
                st.text(f"ID {fi['id']}: {fi['text'][:60]} — {fi['error']}")

    # Translation preview
    with st.expander("📋 Translation preview (first 50)", expanded=False):
        try:
            po_out = polib.pofile(tmp_output, encoding="utf-8")
            preview = []
            count = 0
            for entry in po_out:
                if entry.msgid == "":
                    continue
                preview.append({
                    "msgid": entry.msgid[:80],
                    "msgstr": entry.msgstr[:80] if entry.msgstr else "—",
                })
                count += 1
                if count >= 50:
                    break
            st.dataframe(preview, use_container_width=True, hide_index=True)
        except Exception:
            st.warning("Could not load preview.")

    # Download button
    st.divider()
    try:
        output_bytes = Path(tmp_output).read_bytes()
        stem = Path(original_filename).stem
        download_name = f"translated_{target_lang}_{stem}.po"
        st.download_button(
            label=f"⬇️ Download translated .po",
            data=output_bytes,
            file_name=download_name,
            mime="application/x-gettext",
            type="primary",
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"Could not prepare download: {e}")


def _run_xliff_translation(
    tmp_input, model, batch_size, target_lang, source_lang,
    force, context_text, original_filename, glossary=None,
):
    """Execute XLIFF translation and display results."""

    context_file = None
    if context_text:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".txt", mode="w", encoding="utf-8") as ctx_tmp:
            ctx_tmp.write(context_text)
            context_file = ctx_tmp.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlf") as out_tmp:
        tmp_output = out_tmp.name

    progress_bar = st.progress(0, text="Starting XLIFF translation…")
    log_container = st.empty()
    log_messages = []

    def on_progress(translated, total):
        pct = translated / total if total > 0 else 1.0
        progress_bar.progress(pct, text=f"Translated {translated} / {total} trans-units")

    def on_log(msg):
        log_messages.append(msg)
        log_container.text("\n".join(log_messages[-5:]))

    # XLIFF items with markup content work best at smaller batch sizes
    effective_batch = min(batch_size, 10)

    try:
        result = translate_xliff_file(
            tmp_input,
            tmp_output,
            model=model,
            batch_size=effective_batch,
            target_lang=target_lang,
            source_lang=source_lang,
            force=force,
            context_file=context_file,
            progress_callback=on_progress,
            log_callback=on_log,
            glossary=glossary,
        )
    except Exception as e:
        st.error(f"❌ XLIFF translation failed: {e}")
        return

    progress_bar.progress(1.0, text="✅ Translation complete!")

    st.divider()
    st.subheader("📊 Results")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("✅ Translated", result["translated"])
    col2.metric("📄 Total trans-units", result["total_entries"])
    col3.metric("⚠️ Placeholder issues", len(result["placeholder_warnings"]))
    col4.metric("❌ Failed", len(result["failed"]))

    if result["placeholder_warnings"]:
        with st.expander(f"⚠️ {len(result['placeholder_warnings'])} placeholder warning(s)", expanded=True):
            pw_data = [
                {
                    "ID": pw["id"],
                    "Missing": ", ".join(pw["missing"]),
                    "Source": pw["source"][:80],
                    "Translation": pw["translation"][:80],
                }
                for pw in result["placeholder_warnings"]
            ]
            st.dataframe(pw_data, use_container_width=True, hide_index=True)

    if result["failed"]:
        with st.expander(f"❌ {len(result['failed'])} failed item(s)"):
            for fi in result["failed"]:
                st.text(f"ID {fi['id']}: {fi['text'][:60]} — {fi['error']}")

    # Download
    st.divider()
    try:
        output_bytes = Path(tmp_output).read_bytes()
        stem = Path(original_filename).stem
        download_name = f"translated_{target_lang}_{stem}.xlf"
        st.download_button(
            label="⬇️ Download translated .xlf",
            data=output_bytes,
            file_name=download_name,
            mime="application/xliff+xml",
            type="primary",
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"Could not prepare download: {e}")


if __name__ == "__main__":
    main()
